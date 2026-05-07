"""
설문 통계 엑셀 다운로드 — 예시 형식 생성기

대안 A: 4개 시트 분리
  시트1 설문 요약           (key-value)
  시트2 조직별 응답현황      (raw table)
  시트3 문항별 응답현황      (블록형, 문항 유형별 분기)
  시트4 응답자별 문항답변    (raw data, 동적 컬럼)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ────────────── 스타일 ──────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="305496")  # 진한 파랑
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
KEY_FONT     = Font(bold=True)
KEY_FILL     = PatternFill("solid", fgColor="DDEBF7")  # 연한 파랑
LOW_FILL     = PatternFill("solid", fgColor="D9D9D9")  # 회색 (low rate)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN         = Side(border_style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, max_width=60):
    for col_cells in ws.columns:
        col = col_cells[0].column
        # MergedCell 안전 처리
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).split("\n"):
                # 한글은 폭이 더 넓다고 가정 (영문 1, 한글 2)
                w = sum(2 if ord(ch) > 127 else 1 for ch in line)
                length = max(length, w)
        ws.column_dimensions[get_column_letter(col)].width = min(length + 2, max_width)


# ────────────── 예시 데이터 ──────────────
summary = {
    "surveyId": 1042,
    "title":    "2026년 상반기 임직원 만족도 설문",
    "site":     "본사",
    "beginDate": datetime(2026, 4, 15,  9,  0),
    "endDate":   datetime(2026, 5, 15, 18,  0),
    "totalQuestionCount": 5,
    "totalTargetCount":   1250,
    "respondedCount":     987,
    "notRespondedCount":  263,
    "responseRate":       79.0,
    "daysLeft":           8,
}

dept_rates = [
    # (deptId, deptName, target, responded, rate, lowRate)
    (101, "경영지원본부",   120, 110, 91.7, False),
    (102, "개발1팀",        85,   80, 94.1, False),
    (103, "개발2팀",        92,   58, 63.0, True),   # low
    (104, "CIT LAB팀",      45,   42, 93.3, False),
    (105, "마케팅팀",       60,   38, 63.3, True),   # low
    (106, "영업1팀",       150, 130, 86.7, False),
    (107, "영업2팀",       140,  95, 67.9, True),    # low
    (108, "인프라팀",       38,   36, 94.7, False),
    (109, "디자인팀",       30,   27, 90.0, False),
    (110, "법무팀",         12,   11, 91.7, False),
]

# 문항 메타 (4번 시트 헤더 + 라벨 매핑용)
QUESTIONS = [
    {
        "questionId": 1, "type": "SINGLE_CHOICE",
        "title": "주로 사용하는 OS는?",
        "items": {1: "Windows", 2: "macOS", 3: "Linux"},
    },
    {
        "questionId": 2, "type": "MULTIPLE_CHOICE",
        "title": "최근 1개월 내 사용한 협업 툴은? (복수선택)",
        "items": {11: "Slack", 12: "Notion", 13: "Jira", 14: "Confluence", 15: "Figma"},
    },
    {
        "questionId": 3, "type": "SCALE",
        "title": "회사에 대한 전반적인 만족도는?",
        "items": {21: "1", 22: "2", 23: "3", 24: "4", 25: "5"},
    },
    {
        "questionId": 4, "type": "SUBJECTIVE",
        "title": "회사에 바라는 점이 있다면 자유롭게 작성해주세요.",
        "items": {},
    },
    {
        "questionId": 5, "type": "RANKING",
        "title": "복지제도 중요도 순위를 선택해주세요.",
        "items": {31: "재택근무", 32: "교육비 지원", 33: "건강검진", 34: "리프레시 휴가"},
    },
]

# 문항별 통계 (시트3)
question_stats = [
    {**QUESTIONS[0], "totalResponseCount": 987,
     "stat_items": [("Windows", 612, 62.0), ("macOS", 298, 30.2), ("Linux", 77, 7.8)]},
    {**QUESTIONS[1], "totalResponseCount": 987,
     "stat_items": [("Slack", 901, 91.3), ("Notion", 812, 82.3), ("Jira", 644, 65.2),
                    ("Confluence", 503, 51.0), ("Figma", 287, 29.1)]},
    {**QUESTIONS[2], "totalResponseCount": 987, "average": 4.05,
     "stat_items": [("1", 18, 1.8), ("2", 47, 4.8), ("3", 165, 16.7),
                    ("4", 401, 40.6), ("5", 356, 36.1)]},
    {**QUESTIONS[3], "totalResponseCount": 412, "sampleTexts": [
        "재택근무가 좀 더 유연했으면 좋겠습니다.",
        "교육비 지원 한도를 늘려주세요.",
        "회의가 너무 많습니다. 회의 없는 날을 운영했으면 합니다.",
        "사내 카페 운영 시간 연장 부탁드립니다.",
        "건강검진 항목이 좀 더 다양해졌으면 합니다.",
    ]},
    {**QUESTIONS[4], "totalResponseCount": 987,
     "stat_items": [("재택근무", 421, 42.7), ("교육비 지원", 287, 29.1),
                    ("건강검진", 198, 20.1), ("리프레시 휴가", 81, 8.2)]},
]

# 응답자별 raw 데이터 (시트4) — answers 는 SurveyAnswerDto 형태 가정
respondents = [
    {"empName": "홍석균", "deptName": "CIT LAB팀",
     "submittedAt": datetime(2026, 4, 16, 10, 12),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [2]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11, 12, 15]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 5},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": "전반적으로 만족합니다."},
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [31, 33, 32, 34]},
     ]},
    {"empName": "김지원", "deptName": "개발1팀",
     "submittedAt": datetime(2026, 4, 16, 11, 4),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [1]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11, 13]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 4},
        # 문항4 미응답
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [32, 31, 34, 33]},
     ]},
    {"empName": "박서연", "deptName": "마케팅팀",
     "submittedAt": datetime(2026, 4, 17,  9, 31),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [2]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11, 12, 14, 15]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 3},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": "회의를 줄여주세요."},
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [34, 31, 33, 32]},
     ]},
    {"empName": "이도현", "deptName": "영업1팀",
     "submittedAt": datetime(2026, 4, 17, 14, 22),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [1]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 4},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": ""},  # 빈 답변
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [31, 32, 33, 34]},
     ]},
    {"empName": "최민지", "deptName": "디자인팀",
     "submittedAt": datetime(2026, 4, 18, 16,  3),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [2]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [12, 15]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 5},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": "리프레시 휴가가 잘 활용됐으면 좋겠어요."},
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [34, 33, 31, 32]},
     ]},
    {"empName": "정현우", "deptName": "개발2팀",
     "submittedAt": datetime(2026, 4, 19,  8, 45),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [3]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11, 13, 14]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 2},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": "리눅스 개발 환경 지원이 부족합니다."},
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [32, 33, 31, 34]},
     ]},
    {"empName": "김하은", "deptName": "인프라팀",
     "submittedAt": datetime(2026, 4, 20, 13, 11),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [1]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [13, 14]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 4},
        # 문항4, 문항5 미응답
     ]},
    {"empName": "윤재호", "deptName": "경영지원본부",
     "submittedAt": datetime(2026, 4, 21, 17, 28),
     "answers": [
        {"questionId": 1, "type": "SINGLE_CHOICE",   "selectedItemIds": [1]},
        {"questionId": 2, "type": "MULTIPLE_CHOICE", "selectedItemIds": [11, 12]},
        {"questionId": 3, "type": "SCALE",           "scaleValue": 3},
        {"questionId": 4, "type": "SUBJECTIVE",      "textAnswer": "사내 식당 메뉴가 더 다양했으면 합니다."},
        {"questionId": 5, "type": "RANKING",         "rankedItemIds": [33, 31, 34, 32]},
     ]},
]


# ────────────── 헬퍼 ──────────────
def fmt_dt(dt):
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def answer_to_text(answer, question):
    """SurveyAnswerDto → 사람이 읽을 수 있는 셀 값"""
    t = answer["type"]
    items = question["items"]
    if t == "SINGLE_CHOICE":
        ids = answer.get("selectedItemIds") or []
        return items.get(ids[0], "(삭제됨)") if ids else ""
    if t == "MULTIPLE_CHOICE":
        ids = answer.get("selectedItemIds") or []
        return ", ".join(items.get(i, "(삭제됨)") for i in ids)
    if t == "SUBJECTIVE":
        return answer.get("textAnswer") or ""
    if t == "SCALE":
        v = answer.get("scaleValue")
        return f"{v}점" if v is not None else ""
    if t == "RANKING":
        ids = answer.get("rankedItemIds") or []
        return " / ".join(f"{i+1}.{items.get(qid, '(삭제됨)')}" for i, qid in enumerate(ids))
    return ""


# ────────────── 시트 1: 설문 요약 ──────────────
def build_summary_sheet(wb, s):
    ws = wb.create_sheet("설문 요약")
    rows = [
        ("설문 ID",       s["surveyId"]),
        ("설문명",        s["title"]),
        ("사이트",        s["site"]),
        ("시작일",        fmt_dt(s["beginDate"])),
        ("종료일",        fmt_dt(s["endDate"])),
        ("총 문항 수",    f'{s["totalQuestionCount"]}개'),
        ("응답 대상",     f'{s["totalTargetCount"]:,}명'),
        ("응답 완료",     f'{s["respondedCount"]:,}명'),
        ("미응답",        f'{s["notRespondedCount"]:,}명'),
        ("응답률",        f'{s["responseRate"]}%'),
        ("남은 일수",     f'{s["daysLeft"]}일'),
    ]
    ws["A1"] = "설문 기본정보 요약"
    ws["A1"].font = SECTION_FONT
    ws.merge_cells("A1:B1")

    ws.append(["항목", "값"])
    style_header_row(ws, 2, 2)

    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = KEY_FONT
        ws.cell(row=i, column=1).fill = KEY_FILL
        ws.cell(row=i, column=2, value=v)
        for c in (1, 2):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).alignment = LEFT

    autosize(ws)


# ────────────── 시트 2: 조직별 응답현황 ──────────────
def build_dept_sheet(wb, rows):
    ws = wb.create_sheet("조직별 응답현황")
    headers = ["부서ID", "부서명", "응답대상", "응답완료", "응답률(%)"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for r in rows:
        dept_id, dept_name, target, responded, rate, low = r
        ws.append([dept_id, dept_name, target, responded, rate])
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c != 2 else LEFT
            if low:
                cell.fill = LOW_FILL

    autosize(ws)
    ws.freeze_panes = "A2"


# ────────────── 시트 3: 문항별 응답현황 (블록형) ──────────────
def build_question_sheet(wb, stats):
    ws = wb.create_sheet("문항별 응답현황")
    row = 1

    type_label = {
        "SINGLE_CHOICE": "단일선택", "MULTIPLE_CHOICE": "복수선택",
        "SUBJECTIVE": "주관식", "SCALE": "척도", "RANKING": "순위선택",
    }

    for idx, q in enumerate(stats, start=1):
        # 문항 헤더 (병합)
        title = f'[문항 {idx}] {type_label[q["type"]]} — {q["title"]}'
        meta  = f'응답: {q["totalResponseCount"]:,}건'
        if q["type"] == "SCALE":
            meta += f' / 평균: {q["average"]}점'

        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        ws.cell(row=row, column=1, value=meta).font = Font(italic=True, color="595959")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        if q["type"] == "SUBJECTIVE":
            ws.cell(row=row, column=1, value="No.").fill = HEADER_FILL
            ws.cell(row=row, column=2, value="샘플 응답 텍스트").fill = HEADER_FILL
            for c in (1, 2):
                ws.cell(row=row, column=c).font = HEADER_FONT
                ws.cell(row=row, column=c).alignment = CENTER
                ws.cell(row=row, column=c).border = BORDER
            row += 1
            for i, text in enumerate(q["sampleTexts"], start=1):
                ws.cell(row=row, column=1, value=i).alignment = CENTER
                ws.cell(row=row, column=2, value=text).alignment = LEFT
                for c in (1, 2):
                    ws.cell(row=row, column=c).border = BORDER
                row += 1
        else:
            # 단일/복수/순위/척도 공통: 항목 표
            label_header = "점수" if q["type"] == "SCALE" else "항목"
            count_header = "1순위 응답수" if q["type"] == "RANKING" else "응답수"
            headers = [label_header, count_header, "비율(%)"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
            row += 1
            for label, count, pct in q["stat_items"]:
                ws.cell(row=row, column=1, value=label).alignment = LEFT
                ws.cell(row=row, column=2, value=count).alignment = CENTER
                ws.cell(row=row, column=3, value=pct).alignment   = CENTER
                for c in (1, 2, 3):
                    ws.cell(row=row, column=c).border = BORDER
                row += 1

        row += 1  # 블록 사이 공백

    autosize(ws, max_width=70)


# ────────────── 시트 4: 응답자별 문항답변 ──────────────
def build_respondent_sheet(wb, questions, respondents):
    ws = wb.create_sheet("응답자별 문항답변")

    # 헤더: 고정 3개 + 문항 N개
    fixed = ["응답자명", "부서", "응답일시"]
    q_headers = [f'[문항{i+1}] {q["title"]}' for i, q in enumerate(questions)]
    headers = fixed + q_headers
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    q_by_id = {q["questionId"]: q for q in questions}

    for r in respondents:
        answers_by_qid = {a["questionId"]: a for a in r["answers"]}
        row = [r["empName"], r["deptName"], fmt_dt(r["submittedAt"])]
        for q in questions:
            ans = answers_by_qid.get(q["questionId"])
            row.append(answer_to_text(ans, q) if ans else "")
        ws.append(row)
        row_idx = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c > 3 else CENTER

    # 컬럼 폭: 문항 컬럼은 좀 더 넓게
    autosize(ws, max_width=55)
    for i in range(len(fixed) + 1, len(headers) + 1):
        col = get_column_letter(i)
        cur = ws.column_dimensions[col].width or 10
        ws.column_dimensions[col].width = max(cur, 28)

    ws.freeze_panes = "D2"  # 응답자명/부서/응답일시 + 헤더 고정


# ────────────── 메인 ──────────────
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

build_summary_sheet(wb, summary)
build_dept_sheet(wb, dept_rates)
build_question_sheet(wb, question_stats)
build_respondent_sheet(wb, QUESTIONS, respondents)

# 파일명: 설문통계_{설문명}_{YYYYMMDD-HHmm}.xlsx
fname = f'설문통계_{summary["title"]}_{datetime.now().strftime("%Y%m%d-%H%M")}.xlsx'
out_path = f"/tmp/excel_sample/{fname}"
wb.save(out_path)
print(f"saved: {out_path}")
